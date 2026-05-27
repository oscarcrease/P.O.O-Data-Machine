#Illawarra Data Machine (IDM)


import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import load_workbook


def import_beach_watch_xlsx(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    header_row = None
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row and any(str(x).strip() == "Sample Date" for x in row if x is not None):
            header_row = i
            break

    if header_row is None:
        raise ValueError("Could not find header row containing 'Sample Date'")

    df = pd.read_excel(path, sheet_name=ws.title, header=header_row - 1, engine="openpyxl")
    df = df.dropna(axis=1, how="all").dropna(how="all").reset_index(drop=True)
    df.columns = [str(c).strip() for c in df.columns]

    return df


def clean_beach_watch_df(df):
    df = df.copy()

    rename_map = {
        "Time": "Sample Time",
        "Field Conductivity (mS/cm)": "Conductivity (µS/cm)",
        "Dissolved oxygen (mg/L)": "Dissolved oxygen (mg/L)",
        "No Of Swimmers": "Number of swimmers",
        "Blue Bottles": "Bluebottles",
        "Leaf Litter": "Leaf litter",
        "Marine Debris": "Debris",
        "Surface Foam Scum": "Surface scum",
        "Weeds": "Weed",
        "Enterococci Result (CFU/100mL)": "Enterococci Result",
        "Water Temperature (0C)": "Water Temperature (℃)",
    }
    df = df.rename(columns=rename_map)

    for col in df.columns:
        if df[col].dtype == "object":
            df[col] = df[col].replace(r"^\s*$", np.nan, regex=True)

    key_cols = [c for c in ["Sample Date", "Site Sample Run", "Site", "Sampler"] if c in df.columns]
    if key_cols:
        df = df.dropna(subset=key_cols, how="all").copy()

    if "Sample Date" in df.columns:
        sample_date_dt = pd.to_datetime(df["Sample Date"], errors="coerce", dayfirst=True)
        df["Sample Date"] = sample_date_dt.dt.strftime("%d/%m/%Y")
        df.loc[sample_date_dt.isna(), "Sample Date"] = np.nan

    if "Sample Time" in df.columns:
        t = pd.to_datetime(df["Sample Time"], errors="coerce")
        df["Sample Time"] = t.dt.strftime("%H:%M:%S") + ".000Z"
        df.loc[t.isna(), "Sample Time"] = np.nan

    numeric_cols = [
        "Water Temperature (℃)",
        "Conductivity (µS/cm)",
        "Dissolved oxygen (mg/L)",
        "Number of swimmers",
        "Enterococci Result",
        "Drain Flow",
        "Lagoon Flow",
        "Surface scum",
        "Leaf litter",
        "Litter",
        "Debris",
        "Bluebottles",
        "Weed",
        "Esky Temperature (℃)",
    ]
    for col in numeric_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    if "Sample Run" not in df.columns and "Site Sample Run" in df.columns:
        df["Sample Run"] = df["Site Sample Run"]

    if "Sample Run Submission" not in df.columns:
        df["Sample Run Submission"] = np.nan

    valid_submission = (
        df["Sample Run"].notna() & df["Sample Date"].notna()
        if "Sample Run" in df.columns and "Sample Date" in df.columns
        else pd.Series(False, index=df.index)
    )
    df.loc[valid_submission, "Sample Run Submission"] = (
        df.loc[valid_submission, "Sample Run"].astype(str)
        + "-"
        + df.loc[valid_submission, "Sample Date"].astype(str)
    )

    df["Status"] = np.nan
    real_row_mask = pd.Series(False, index=df.index)
    for c in ["Sample Run", "Sample Date", "Site", "Site Sample Run"]:
        if c in df.columns:
            real_row_mask = real_row_mask | df[c].notna()
    df.loc[real_row_mask, "Status"] = "Awaiting Approval"

    df["No Enterococci Result"] = np.nan
    if "Enterococci Result" in df.columns:
        real_sample_mask = pd.Series(False, index=df.index)

        for c in ["Site", "Site Sample Run", "Sample Run", "Sample Date"]:
            if c in df.columns:
                real_sample_mask = real_sample_mask | df[c].notna()

        no_enterococci_mask = df["Enterococci Result"].isna() & real_sample_mask
        df.loc[no_enterococci_mask, "No Enterococci Result"] = 1

    desired_order = [
        "Sample Run",
        "Sample Run Submission",
        "Sample Date",
        "Sampler",
        "Observation Type",
        "Site Sample Run",
        "Site",
        "Other Site Name",
        "Sample Time",
        "Weather",
        "Drain Flow",
        "Lagoon Flow",
        "Water Temperature (℃)",
        "Conductivity (µS/cm)",
        "Dissolved oxygen (mg/L)",
        "Number of swimmers",
        "Surface scum",
        "Leaf litter",
        "Litter",
        "Debris",
        "Bluebottles",
        "Weed",
        "Visual Turbidity",
        "Esky Temperature (℃)",
        "Comments",
        "Enterococci Result",
        "No Enterococci Result",
        "Status",
    ]

    for col in desired_order:
        if col not in df.columns:
            df[col] = np.nan

    df = df[desired_order].reset_index(drop=True)

    return df


def export_beach_watch_csv(df, output_dir="output"):
    df = df.copy()

    if "Sample Date" not in df.columns:
        raise ValueError("Sample Date column not found")

    sample_dates = pd.to_datetime(df["Sample Date"], format="%d/%m/%Y", errors="coerce")
    valid_dates = sample_dates.dropna()

    if valid_dates.empty:
        raise ValueError("No valid dates found in Sample Date column")

    file_date = valid_dates.min().strftime("%d %B %Y")
    filename = f"{file_date} Illawarra Beaches Data.csv"

    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    full_path = output_path / filename

    if full_path.exists():
        try:
            full_path.unlink()
        except PermissionError:
            stem = f"{file_date} Illawarra Beaches Data"
            i = 1
            while True:
                alt = output_path / f"{stem} ({i}).csv"
                if not alt.exists():
                    full_path = alt
                    break
                i += 1

    df.to_csv(full_path, index=False, encoding="utf-8-sig")
    return full_path


raw_df = import_beach_watch_xlsx(
    r"C:\Users\creaso01\PycharmProjects\Illawarra_Data_Machine\Beach Watch.xlsx"
)

clean_df = clean_beach_watch_df(raw_df)

print(clean_df.head())
print(clean_df.dtypes)

csv_path = export_beach_watch_csv(clean_df)
print(csv_path)

csv_path = export_beach_watch_csv(clean_df)
print(csv_path)

