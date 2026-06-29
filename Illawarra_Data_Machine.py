import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from openpyxl import load_workbook

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, LongTable, TableStyle, Paragraph, Spacer
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak

st.set_page_config(
    page_title="Illawarra Data Machine",
    page_icon="🌊",
    layout="wide"
)

st.markdown(
    """
    <style>
    .block-container {padding-top: 2rem; padding-bottom: 2rem;}
    .idm-card {
        background: #f7f9fc;
        border: 1px solid #dbe4f0;
        border-radius: 14px;
        padding: 1rem 1.2rem;
        margin-bottom: 1rem;
    }
    .idm-small {color: #5b6573; font-size: 0.95rem;}
    .idm-title {font-size: 2rem; font-weight: 700; margin-bottom: 0.25rem;}
    .idm-pill {
        display: inline-block;
        background: #e8f2ff;
        color: #0b5cab;
        padding: 0.2rem 0.6rem;
        border-radius: 999px;
        font-size: 0.85rem;
        font-weight: 600;
        margin-bottom: 0.6rem;
    }
    </style>
    """,
    unsafe_allow_html=True,
)


def import_beach_watch_xlsx(uploaded_file):
    uploaded_file.seek(0)
    wb = load_workbook(uploaded_file, read_only=True, data_only=True)
    ws = wb.active

    header_row = None
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row and any(str(x).strip() == "Sample Date" for x in row if x is not None):
            header_row = i
            break

    if header_row is None:
        raise ValueError("Could not find header row containing 'Sample Date'")

    uploaded_file.seek(0)
    df = pd.read_excel(uploaded_file, sheet_name=ws.title, header=header_row - 1, engine="openpyxl")
    df = df.dropna(axis=1, how="all").dropna(how="all").reset_index(drop=True)
    df.columns = [str(c).strip() for c in df.columns]
    return df


def clean_beach_watch_df(df):
    df = df.copy()

    for col in ["Sample Run", "Sample Run Submission", "Site Sample Run", "Site", "Sampler", "Status"]:
        if col in df.columns:
            df[col] = df[col].astype("object")

    rename_map = {
        "Time": "Sample Time",
        "Field Conductivity (mS/cm)": "Conductivity (µS/cm)",
        "Dissolved oxygen (mg/L)": "Dissolved oxygen (mg/L)",
        "No Of Swimmers": "Number of swimmers",
        "Blue Bottles": "Bluebottles",
        "Leaf Litter": "Leaf litter",
        "Marine Debris": "Debris",
        "Surface Foam Scum": "Surface scum",
        "Weeds": "Weed",
        "Enterococci Result (CFU/100mL)": "Enterococci Result",
        "Water Temperature (0C)": "Water Temperature (℃)",
    }
    df = df.rename(columns=rename_map)

    for col in df.columns:
        if df[col].dtype == "object":
            df[col] = df[col].replace(r"^\s*$", np.nan, regex=True)

    key_cols = [c for c in ["Sample Date", "Site Sample Run", "Site", "Sampler"] if c in df.columns]
    if key_cols:
        df = df.dropna(subset=key_cols, how="all").copy()

    if "Sample Date" in df.columns:
        sample_date_dt = pd.to_datetime(df["Sample Date"], errors="coerce", dayfirst=True)
        df["Sample Date"] = sample_date_dt.apply(
            lambda x: f"{x.day}/{x.month}/{x.year}" if pd.notna(x) else np.nan
        )

    if "Sample Time" in df.columns:
        t = pd.to_datetime(df["Sample Time"], errors="coerce")
        df["Sample Time"] = t.dt.strftime("%H:%M:%S") + ".000Z"
        df.loc[t.isna(), "Sample Time"] = np.nan

    numeric_cols = [
        "Water Temperature (℃)",
        "Conductivity (µS/cm)",
        "Dissolved oxygen (mg/L)",
        "Number of swimmers",
        "Enterococci Result",
        "Drain Flow",
        "Lagoon Flow",
        "Surface scum",
        "Leaf litter",
        "Litter",
        "Debris",
        "Bluebottles",
        "Weed",
        "Esky Temperature (℃)",
    ]
    for col in numeric_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    if "Conductivity (µS/cm)" in df.columns:
        df["Conductivity (µS/cm)"] = df["Conductivity (µS/cm)"] * 1000

    real_sample_mask = pd.Series(False, index=df.index)
    for c in ["Site", "Site Sample Run", "Sample Date", "Sampler"]:
        if c in df.columns:
            real_sample_mask = real_sample_mask | df[c].notna()

    df["Sample Run"] = pd.Series(np.nan, index=df.index, dtype="object")
    df.loc[real_sample_mask, "Sample Run"] = "Illawarra Beaches"

    df["Sample Run Submission"] = pd.Series(np.nan, index=df.index, dtype="object")
    if "Sample Date" in df.columns:
        valid_submission = real_sample_mask & df["Sample Date"].notna()
        df.loc[valid_submission, "Sample Run Submission"] = (
            "Illawarra Beaches-"
            + df.loc[valid_submission, "Sample Date"].astype(str)
        )

    df["Status"] = pd.Series(np.nan, index=df.index, dtype="object")
    real_row_mask = pd.Series(False, index=df.index)
    for c in ["Sample Run", "Sample Date", "Site", "Site Sample Run"]:
        if c in df.columns:
            real_row_mask = real_row_mask | df[c].notna()
    df.loc[real_row_mask, "Status"] = "Awaiting Approval"

    df["No Enterococci Result"] = np.nan
    if "Enterococci Result" in df.columns:
        real_sample_mask = pd.Series(False, index=df.index)
        for c in ["Site", "Site Sample Run", "Sample Run", "Sample Date"]:
            if c in df.columns:
                real_sample_mask = real_sample_mask | df[c].notna()
        no_enterococci_mask = df["Enterococci Result"].isna() & real_sample_mask
        df.loc[no_enterococci_mask, "No Enterococci Result"] = 1

    desired_order = [
        "Sample Run",
        "Sample Run Submission",
        "Sample Date",
        "Sampler",
        "Observation Type",
        "Site Sample Run",
        "Site",
        "Other Site Name",
        "Sample Time",
        "Weather",
        "Drain Flow",
        "Lagoon Flow",
        "Water Temperature (℃)",
        "Conductivity (µS/cm)",
        "Dissolved oxygen (mg/L)",
        "Number of swimmers",
        "Surface scum",
        "Leaf litter",
        "Litter",
        "Debris",
        "Bluebottles",
        "Weed",
        "Visual Turbidity",
        "Esky Temperature (℃)",
        "Comments",
        "Enterococci Result",
        "No Enterococci Result",
        "Status",
    ]

    for col in desired_order:
        if col not in df.columns:
            if col in ["Sample Run", "Sample Run Submission", "Status"]:
                df[col] = pd.Series(np.nan, index=df.index, dtype="object")
            else:
                df[col] = np.nan

    export_key_cols = [
        "Sample Run Submission",
        "Sample Date",
        "Site Sample Run",
        "Site",
        "Sample Time",
    ]

    existing_export_key_cols = [col for col in export_key_cols if col in df.columns]
    if existing_export_key_cols:
        df = df.dropna(subset=existing_export_key_cols, how="all").copy()

    return df[desired_order].reset_index(drop=True)


def get_output_filename(df):
    if "Sample Date" not in df.columns:
        raise ValueError("Sample Date column not found")

    sample_dates = pd.to_datetime(df["Sample Date"], errors="coerce", dayfirst=True)
    valid_dates = sample_dates.dropna()

    if valid_dates.empty:
        raise ValueError("No valid dates found in Sample Date column")

    file_date = valid_dates.min().strftime("%d %B %Y")
    return f"{file_date} Illawarra Beaches Data.csv"


def dataframe_to_csv_bytes(df):
    return df.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")

def reorder_checker_df(df):
    df = df.copy()

    remove_cols = [
        "Sample Run",
        "Sample Run Submission",
        "Sampler",
        "Observation Type",
        "Site Sample Run",
        "Other Site Name",
        "Status",
    ]

    existing_remove_cols = [col for col in remove_cols if col in df.columns]
    if existing_remove_cols:
        df = df.drop(columns=existing_remove_cols)

    if "Site" in df.columns and "Enterococci Result" in df.columns:
        cols = df.columns.tolist()
        cols.remove("Enterococci Result")
        site_idx = cols.index("Site")
        cols.insert(site_idx + 1, "Enterococci Result")
        df = df[cols]

    sort_cols = [col for col in ["Site", "Sample Date", "Sample Time"] if col in df.columns]
    if sort_cols:
        df = df.sort_values(by=sort_cols, ascending=True, na_position="last").reset_index(drop=True)

    return df


def get_pdf_filename(df):
    if "Sample Date" not in df.columns:
        raise ValueError("Sample Date column not found")

    sample_dates = pd.to_datetime(df["Sample Date"], errors="coerce", dayfirst=True)
    valid_dates = sample_dates.dropna()

    if valid_dates.empty:
        raise ValueError("No valid dates found in Sample Date column")

    file_date = valid_dates.min().strftime("%d %B %Y")
    return f"{file_date} Illawarra Beaches data checker.pdf"


def _format_pdf_value(val):
    if pd.isna(val):
        return ""
    if isinstance(val, (float, np.floating)):
        if np.isfinite(val):
            if float(val).is_integer():
                return str(int(val))
            return f"{val:g}"
        return ""
    return str(val)


def dataframe_to_pdf_bytes(df, title):
    pdf_buffer = BytesIO()

    page_width, page_height = landscape(A4)
    left_margin = 10 * mm
    right_margin = 10 * mm
    top_margin = 12 * mm
    bottom_margin = 12 * mm
    usable_width = page_width - left_margin - right_margin

    doc = SimpleDocTemplate(
        pdf_buffer,
        pagesize=landscape(A4),
        leftMargin=left_margin,
        rightMargin=right_margin,
        topMargin=top_margin,
        bottomMargin=bottom_margin,
    )

    styles = getSampleStyleSheet()

    title_style = styles["Title"].clone("checker_title")
    title_style.fontName = "Helvetica-Bold"
    title_style.fontSize = 18
    title_style.leading = 22
    title_style.spaceAfter = 6

    header_style = styles["BodyText"].clone("header_style")
    header_style.fontName = "Helvetica-Bold"
    header_style.fontSize = 7
    header_style.leading = 8.5
    header_style.textColor = colors.white
    header_style.alignment = 1
    header_style.splitLongWords = 0
    header_style.wordWrap = "CJK"
    header_style.spaceAfter = 0
    header_style.spaceBefore = 0

    body_style = styles["BodyText"].clone("body_style")
    body_style.fontName = "Helvetica"
    body_style.fontSize = 7.5
    body_style.leading = 9
    body_style.splitLongWords = 0
    body_style.wordWrap = "CJK"
    body_style.spaceAfter = 0
    body_style.spaceBefore = 0

    site_style = styles["BodyText"].clone("site_style")
    site_style.fontName = "Helvetica-Bold"
    site_style.fontSize = 12
    site_style.leading = 14
    site_style.splitLongWords = 0
    site_style.wordWrap = None
    site_style.spaceAfter = 0
    site_style.spaceBefore = 0

    entero_style = styles["BodyText"].clone("entero_style")
    entero_style.fontName = "Helvetica-Bold"
    entero_style.fontSize = 13
    entero_style.leading = 15
    entero_style.alignment = 1
    entero_style.splitLongWords = 0
    entero_style.wordWrap = "CJK"
    entero_style.spaceAfter = 0
    entero_style.spaceBefore = 0

    def fmt(val):
        if pd.isna(val):
            return ""
        if isinstance(val, (float, np.floating)):
            if np.isfinite(val):
                if float(val).is_integer():
                    return str(int(val))
                return f"{val:g}"
            return ""
        return str(val)

    pdf_df = df.copy()

    preferred_order = [
        "Sample Date",
        "Site",
        "Enterococci Result",
        "Sample Time",
        "Weather",
        "Drain Flow",
        "Lagoon Flow",
        "Water Temperature (℃)",
        "Conductivity (µS/cm)",
        "Dissolved oxygen (mg/L)",
        "Number of swimmers",
        "Surface scum",
        "Leaf litter",
        "Litter",
        "Debris",
        "Bluebottles",
        "Weed",
        "Visual Turbidity",
        "Esky Temperature (℃)",
        "Comments",
        "No Enterococci Result",
    ]
    existing_cols = [c for c in preferred_order if c in pdf_df.columns]
    pdf_df = pdf_df[existing_cols]

    display_headers = {
        "Sample Date": "Sample\nDate",
        "Site": "Site",
        "Enterococci Result": "Enterococci\nResult",
        "Sample Time": "Sample\nTime",
        "Weather": "Weather",
        "Drain Flow": "Drain\nFlow",
        "Lagoon Flow": "Lagoon\nFlow",
        "Water Temperature (℃)": "Water\nTemp (℃)",
        "Conductivity (µS/cm)": "Conductivity\n(µS/cm)",
        "Dissolved oxygen (mg/L)": "Dissolved\nOxygen\n(mg/L)",
        "Number of swimmers": "Number of\nSwimmers",
        "Surface scum": "Surface\nScum",
        "Leaf litter": "Leaf\nLitter",
        "Litter": "Litter",
        "Debris": "Debris",
        "Bluebottles": "Blue-\nbottles",
        "Weed": "Weed",
        "Visual Turbidity": "Visual\nTurbidity",
        "Esky Temperature (℃)": "Esky\nTemp (℃)",
        "Comments": "Comments",
        "No Enterococci Result": "No\nEnterococci\nResult",
    }

    col_width_map = {
        "Sample Date": 20 * mm,
        "Site": 46 * mm,
        "Enterococci Result": 20 * mm,
        "Sample Time": 17 * mm,
        "Weather": 20 * mm,
        "Drain Flow": 10 * mm,
        "Lagoon Flow": 10 * mm,
        "Water Temperature (℃)": 14 * mm,
        "Conductivity (µS/cm)": 16 * mm,
        "Dissolved oxygen (mg/L)": 14 * mm,
        "Number of swimmers": 14 * mm,
        "Surface scum": 10 * mm,
        "Leaf litter": 10 * mm,
        "Litter": 8 * mm,
        "Debris": 8 * mm,
        "Bluebottles": 10 * mm,
        "Weed": 8 * mm,
        "Visual Turbidity": 14 * mm,
        "Esky Temperature (℃)": 14 * mm,
        "Comments": 18 * mm,
        "No Enterococci Result": 14 * mm,
    }

    col_widths = [col_width_map.get(col, 18 * mm) for col in pdf_df.columns]
    total_width = sum(col_widths)
    if total_width > usable_width:
        scale = usable_width / total_width
        col_widths = [w * scale for w in col_widths]

    rows_per_page = 5
    story = []

    for start in range(0, len(pdf_df), rows_per_page):
        chunk = pdf_df.iloc[start:start + rows_per_page]

        if start > 0:
            story.append(PageBreak())

        story.append(Paragraph(title, title_style))
        story.append(Spacer(1, 4 * mm))

        header_row = [
            Paragraph(display_headers.get(str(col), str(col)), header_style)
            for col in pdf_df.columns
        ]
        table_data = [header_row]

        for _, row in chunk.iterrows():
            row_cells = []
            for col, val in row.items():
                text = fmt(val)

                if col == "Site":
                    cell = Paragraph(text, site_style)
                elif col == "Enterococci Result":
                    cell = Paragraph(text, entero_style)
                else:
                    cell = Paragraph(text, body_style)

                row_cells.append(cell)

            table_data.append(row_cells)

        row_heights = [18 * mm] + [None] * len(chunk)

        table = Table(
            table_data,
            colWidths=col_widths,
            rowHeights=row_heights,
            repeatRows=1
        )

        table_styles = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0b5cab")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("ALIGN", (0, 1), (-1, -1), "LEFT"),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#b8c7db")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f4f7fb")]),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, 0), 5),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 5),
            ("TOPPADDING", (0, 1), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 1), (-1, -1), 6),
        ]

        if "Enterococci Result" in pdf_df.columns:
            entero_col = pdf_df.columns.get_loc("Enterococci Result")
            table_styles.append(("ALIGN", (entero_col, 1), (entero_col, -1), "CENTER"))

        table.setStyle(TableStyle(table_styles))
        story.append(table)

    doc.build(story)
    pdf_buffer.seek(0)
    return pdf_buffer.getvalue()

st.markdown('<div class="idm-title"> P.O.Os Illawarra Data Machine</div>', unsafe_allow_html=True)
st.markdown('<div class="idm-small">Automatically cleans Illawarra data, ready for Salesforce Upload.</div>', unsafe_allow_html=True)

with st.sidebar:
    st.header("How it works")
    st.markdown(
        """
1. Upload the raw Illawarra xlsx file.
2. Check that everything looks right in the preview.
3. Hit Download Cleaned CSV, then import that csv into Salesforce using the below steps.
4. Click Download Data Checker PDF to download an easy to review PDF. Attach this to the run in salesforce!
        """
    )
    st.info("""
For Salesforce imports  
1. Go to Submissions  
2. Import
3. Click submissions
4. Click Add new records
5. Change the 4th box down to Site Sample Run Name  
6. Skip 2 boxes  
7. Change the next 3 boxes to Site name, Contact name & Sample Run Submission Name  
8. Drag the downloaded clean CSV and hit next in the bottom right corner
9. Click next again
10. Click start Import! 

Check the sample run submission in salesforce to ensure it imported correctly, and add the checker PDF while you are there!
    """)
        with st.expander("Salesforce Import Screenshot"):
        st.image("assets/SF_import_tutorial.png", caption="Salesforce import field mapping", use_container_width=True)

uploaded_file = st.file_uploader("Upload Beach Watch Excel file", type=["xlsx"])

if uploaded_file is None:
    st.markdown('<div class="idm-card"><strong>Start here:</strong> upload an Excel file above to generate the cleaned CSV.</div>', unsafe_allow_html=True)
else:
    try:
        with st.spinner("Importing and cleaning file..."):
            raw_df = import_beach_watch_xlsx(uploaded_file)
            clean_df = clean_beach_watch_df(raw_df)

            output_filename = get_output_filename(clean_df)
            csv_bytes = dataframe_to_csv_bytes(clean_df)

            checker_df = reorder_checker_df(clean_df)
            checker_pdf_filename = get_pdf_filename(clean_df)
            checker_pdf_bytes = dataframe_to_pdf_bytes(
                checker_df,
                title=checker_pdf_filename.replace(".pdf", "")
            )

        col1, col2, col3 = st.columns(3)
        col1.metric("Rows", f"{len(clean_df):,}")
        col2.metric("Columns", f"{len(clean_df.columns):,}")
        col3.metric("Output file", output_filename)

        dl1, dl2 = st.columns(2)

        dl1.download_button(
            label="Download Cleaned CSV",
            data=csv_bytes,
            file_name=output_filename,
            mime="text/csv",
            use_container_width=True,
        )

        dl2.download_button(
            label="Download Data Checker PDF",
            data=checker_pdf_bytes,
            file_name=checker_pdf_filename,
            mime="application/pdf",
            use_container_width=True,
        )

        with st.expander("Preview cleaned data", expanded=True):
            st.dataframe(clean_df, use_container_width=True, height=420)

        with st.expander("Column names"):
            st.write(clean_df.columns.tolist())

        with st.expander("Raw data preview"):
            st.dataframe(raw_df.head(20), use_container_width=True)

    except Exception as e:
        st.error(f"Error: {e}")
